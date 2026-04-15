import sys
import os
import shutil
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QLineEdit, 
                             QTextEdit, QProgressBar, QFileDialog, QMessageBox,
                             QGroupBox, QGridLayout, QTableWidget, QTableWidgetItem,
                             QHeaderView, QAbstractItemView, QSplitter)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QBrush, QTextCursor
import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.copier import WorksheetCopy

class AnalysisThread(QThread):
    """分析线程 - 避免界面卡死"""
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    status_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str, int)
    result_signal = pyqtSignal(list)
    
    def __init__(self, file_path):
        super().__init__()
        self.original_file_path = file_path
        self.converted_file_path = None
        
    def run(self):
        """线程运行函数"""
        try:
            self.log_signal.emit("=" * 80)
            self.log_signal.emit(f"开始分析文件: {os.path.basename(self.original_file_path)}")
            
            # ============ 1. 文件处理 ============
            file_to_analyze = self.original_file_path
            
            if file_to_analyze.lower().endswith('.xls'):
                self.log_signal.emit("检测到.xls文件，正在转换格式...")
                self.status_signal.emit("正在转换文件格式...")
                
                # 在同目录下创建.xlsx文件
                file_dir = os.path.dirname(self.original_file_path)
                file_name = os.path.splitext(os.path.basename(self.original_file_path))[0]
                xlsx_path = os.path.join(file_dir, f"{file_name}_converted.xlsx")
                
                # 转换文件
                file_to_analyze = self.convert_xls_to_xlsx(self.original_file_path, xlsx_path)
                self.converted_file_path = file_to_analyze
                self.log_signal.emit(f"✓ 转换成功: {os.path.basename(file_to_analyze)}")
                self.log_signal.emit(f"⚠️ 注意：转换过程会丢失公式，仅保留数值")
            
            # ============ 2. 分析CPK文件 ============
            self.log_signal.emit("\n开始分析CAV工作表...")
            self.status_signal.emit("正在分析CPK数据...")
            
            failure_data = self.analyze_cpk_file(file_to_analyze)
            
            # ============ 3. 生成汇总报告 ============
            if failure_data:
                self.log_signal.emit(f"\n✓ 发现 {len(failure_data)} 条CPK失败记录")
                
                # 直接在原文件最后添加汇总表
                output_path = self.add_summary_to_original(file_to_analyze, failure_data)
                self.log_signal.emit(f"✓ 失败记录汇总已添加到文件末尾: {os.path.basename(output_path)}")
                self.log_signal.emit(f"  - 工作表位置: 最后")
                self.log_signal.emit(f"  - 工作表名称: CPK_FAIL_Summary")
                self.log_signal.emit(f"  - 记录数: {len(failure_data)}")
            else:
                self.log_signal.emit("\n✓ 未发现CPK失败记录")
            
            # 发送结果到界面
            self.result_signal.emit(failure_data)
            
            # 完成信号
            self.finished_signal.emit(True, "分析完成", len(failure_data))
            
        except Exception as e:
            self.error_signal.emit(str(e))
            self.finished_signal.emit(False, f"分析失败: {str(e)}", 0)
    
    def convert_xls_to_xlsx(self, xls_path, xlsx_path):
        """
        将.xls转换为.xlsx
        注意：此方法会丢失公式、格式等，仅保留数值
        """
        try:
            if os.path.exists(xlsx_path):
                os.remove(xlsx_path)
            
            self.log_signal.emit("正在读取.xls文件...")
            workbook_xls = xlrd.open_workbook(xls_path, formatting_info=False)
            
            workbook_xlsx = Workbook()
            
            # 移除默认的sheet
            default_sheet = workbook_xlsx.active
            workbook_xlsx.remove(default_sheet)
            
            for sheet_index, sheet_name in enumerate(workbook_xls.sheet_names()):
                self.log_signal.emit(f"  转换工作表: {sheet_name}")
                sheet_xls = workbook_xls.sheet_by_name(sheet_name)
                sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_name[:31])
                
                for row_idx in range(sheet_xls.nrows):
                    for col_idx in range(sheet_xls.ncols):
                        cell_value = sheet_xls.cell_value(row_idx, col_idx)
                        cell = sheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
            
            workbook_xlsx.save(xlsx_path)
            return xlsx_path
            
        except Exception as e:
            raise Exception(f"文件转换失败: {str(e)}")
    
    def excel_date_to_datetime(self, excel_date):
        """将Excel日期数字转换为日期字符串"""
        try:
            if excel_date is None:
                return ""
            
            if isinstance(excel_date, datetime):
                return excel_date.strftime('%Y-%m-%d')
            
            if isinstance(excel_date, (int, float)):
                base_date = datetime(1899, 12, 30)
                converted_date = base_date + timedelta(days=excel_date)
                return converted_date.strftime('%Y-%m-%d')
            
            if isinstance(excel_date, str):
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']:
                    try:
                        return datetime.strptime(excel_date, fmt).strftime('%Y-%m-%d')
                    except:
                        continue
            
            return str(excel_date)
            
        except Exception:
            return str(excel_date) if excel_date else ""
    
    def analyze_cpk_file(self, file_path):
        """分析CPK文件 - 只处理CAV开头的sheet"""
        failure_data = []
        
        try:
            workbook = load_workbook(filename=file_path, data_only=True)
            
            # 只处理以CAV开头的sheet
            cav_sheets = []
            for sheet_name in workbook.sheetnames:
                if sheet_name.strip().upper().startswith('CAV'):
                    cav_sheets.append(sheet_name)
            
            if not cav_sheets:
                self.log_signal.emit("⚠️ 未找到以CAV开头的sheet！")
                return failure_data
            
            self.log_signal.emit(f"找到 {len(cav_sheets)} 个CAV开头的sheet")
            
            total_checked = 0
            total_failures = 0
            
            for sheet_name in cav_sheets:
                sheet = workbook[sheet_name]
                self.log_signal.emit(f"\n▶ 处理sheet: {sheet_name}")
                
                # 获取Date信息 (H6单元格)
                date_cell = sheet['H6']
                date_value = date_cell.value
                date_str = self.excel_date_to_datetime(date_value)
                
                row = 11
                sheet_failures = 0
                
                while True:
                    spc_cell = sheet.cell(row=row, column=2)
                    if not spc_cell.value:
                        break
                    
                    # 获取A列的值（No.）
                    no_cell = sheet.cell(row=row, column=1)
                    no_value = str(no_cell.value) if no_cell.value else ""
                    
                    spc_value = str(spc_cell.value) if spc_cell.value else ""
                    fai_cell = sheet.cell(row=row, column=3)
                    fai_value = str(fai_cell.value) if fai_cell.value else ""
                    
                    target_cell = sheet.cell(row=row, column=5)
                    target_value = target_cell.value
                    
                    cal_cell = sheet.cell(row=row, column=24)
                    cal_value = cal_cell.value
                    
                    target_float = None
                    cal_float = None
                    
                    try:
                        if target_value is not None:
                            target_float = float(target_value)
                        if cal_value is not None:
                            cal_float = float(cal_value)
                    except (ValueError, TypeError):
                        pass
                    
                    if (target_float is not None and cal_float is not None and 
                        cal_float < target_float):
                        
                        failure_record = {
                            'date': date_str,
                            'no': no_value,  # 新增No.字段
                            'cav_no': sheet_name.strip(),
                            'spc': spc_value,
                            'fai': fai_value,
                            'target_cpk': target_float,
                            'cal_cpk': cal_float,
                            'result': 'FAIL',
                            'row': row
                        }
                        failure_data.append(failure_record)
                        sheet_failures += 1
                        total_failures += 1
                        
                        self.log_signal.emit(
                            f"  发现失败: 行{row} | No:{no_value} | "
                            f"Target:{target_float:.3f} | Cal:{cal_float:.3f}"
                        )
                    
                    total_checked += 1
                    row += 1
                
                self.log_signal.emit(f"  {sheet_name}: 检查{row-11}行，发现{sheet_failures}条失败记录")
            
            self.log_signal.emit(f"\n" + "=" * 80)
            self.log_signal.emit(f"分析完成！总计检查 {total_checked} 条记录")
            self.log_signal.emit(f"发现 {total_failures} 条CPK失败记录")
            
            return failure_data
            
        except Exception as e:
            raise Exception(f"分析文件时出错: {str(e)}")
    
    def add_summary_to_original(self, file_path, failure_data):
        """
        直接在原文件最后添加汇总表
        不修改任何原有内容，只在最后添加新sheet
        """
        try:
            # 加载工作簿
            workbook = load_workbook(filename=file_path)
            
            # 删除已有的CPK_FAIL_Summary sheet（如果存在）
            if "CPK_FAIL_Summary" in workbook.sheetnames:
                std = workbook["CPK_FAIL_Summary"]
                workbook.remove(std)
            
            # 创建新的汇总sheet（自动添加到最后）
            summary_sheet = workbook.create_sheet("CPK_FAIL_Summary")
            
            # ============ 设置标题行 ============
            headers = ['Date', 'No.', 'Cav No.', 'SPC', 'FAI#', 'TargetCPK', 'CalCPK', 'Result']
            
            header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col, header in enumerate(headers, 1):
                cell = summary_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # ============ 写入数据 ============
            fail_font = Font(name='微软雅黑', size=10, bold=True, color="FFFFFF")
            fail_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            fail_alignment = Alignment(horizontal='center', vertical='center')
            
            normal_font = Font(name='微软雅黑', size=10)
            normal_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            right_alignment = Alignment(horizontal='right', vertical='center')
            
            for row_idx, record in enumerate(failure_data, 2):
                # Date
                date_cell = summary_sheet.cell(row=row_idx, column=1, value=record['date'])
                date_cell.font = normal_font
                date_cell.alignment = normal_alignment
                date_cell.number_format = 'yyyy-mm-dd'
                
                # No. (新增列)
                no_cell = summary_sheet.cell(row=row_idx, column=2, value=record['no'])
                no_cell.font = normal_font
                no_cell.alignment = normal_alignment
                
                # Cav No.
                cav_cell = summary_sheet.cell(row=row_idx, column=3, value=record['cav_no'])
                cav_cell.font = normal_font
                cav_cell.alignment = normal_alignment
                
                # SPC
                spc_cell = summary_sheet.cell(row=row_idx, column=4, value=record['spc'])
                spc_cell.font = normal_font
                spc_cell.alignment = left_alignment
                
                # FAI#
                fai_cell = summary_sheet.cell(row=row_idx, column=5, value=record['fai'])
                fai_cell.font = normal_font
                fai_cell.alignment = normal_alignment
                
                # TargetCPK
                target_cell = summary_sheet.cell(row=row_idx, column=6, value=record['target_cpk'])
                target_cell.font = normal_font
                target_cell.alignment = right_alignment
                target_cell.number_format = '0.00'
                
                # CalCPK
                cal_cell = summary_sheet.cell(row=row_idx, column=7, value=record['cal_cpk'])
                cal_cell.font = normal_font
                cal_cell.alignment = right_alignment
                cal_cell.number_format = '0.000000'
                
                # Result
                result_cell = summary_sheet.cell(row=row_idx, column=8, value="FAIL")
                result_cell.font = fail_font
                result_cell.fill = fail_fill
                result_cell.alignment = fail_alignment
            
            # ============ 设置列宽 ============
            column_widths = {
                'A': 15,   # Date
                'B': 10,   # No.
                'C': 12,   # Cav No.
                'D': 50,   # SPC
                'E': 12,   # FAI#
                'F': 12,   # TargetCPK
                'G': 15,   # CalCPK
                'H': 10    # Result
            }
            
            for col, width in column_widths.items():
                summary_sheet.column_dimensions[col].width = width
            
            # 设置D列自动换行
            summary_sheet.column_dimensions['D'].alignment = Alignment(wrap_text=True)
            
            # 添加筛选器
            summary_sheet.auto_filter.ref = summary_sheet.dimensions
            
            # 冻结首行
            summary_sheet.freeze_panes = 'A2'
            
            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in summary_sheet.iter_rows(min_row=1, max_row=len(failure_data)+1, max_col=8):
                for cell in row:
                    cell.border = thin_border
            
            # 保存文件
            workbook.save(file_path)
            
            return file_path
            
        except Exception as e:
            raise Exception(f"添加汇总表时出错: {str(e)}")


class CPKAnalyzerWindow(QMainWindow):
    """主窗口"""
    def __init__(self):
        super().__init__()
        self.analysis_thread = None
        self.current_file_path = None
        self.converted_file_path = None
        self.failure_data = []
        self.init_ui()
        
    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle("CPK失败记录分析工具 v4.2")
        self.setGeometry(100, 100, 1300, 850)
        
        # 设置全局样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f7fa;
            }
            QGroupBox {
                font-weight: bold;
                font-size: 11pt;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 10px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 10px 0 10px;
                color: #1976D2;
            }
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 10pt;
            }
            QPushButton:hover {
                background-color: #1565C0;
            }
            QPushButton#danger {
                background-color: #d32f2f;
            }
            QPushButton#success {
                background-color: #2E7D32;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                background-color: white;
                font-size: 10pt;
            }
            QTextEdit {
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 10pt;
                background-color: white;
            }
            QProgressBar {
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                text-align: center;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #2E7D32;
                border-radius: 4px;
            }
            QTableWidget {
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                background-color: white;
                font-size: 9pt;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #2C3E50;
                color: white;
                padding: 8px;
                border: 1px solid #34495e;
                font-weight: bold;
            }
        """)
        
        # 中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # ============ 文件选择区域 ============
        file_group = QGroupBox("1. 文件选择")
        file_layout = QGridLayout(file_group)
        file_layout.setVerticalSpacing(10)
        
        file_layout.addWidget(QLabel("原始文件："), 0, 0)
        
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("请选择CPK报告文件 (.xls / .xlsx)...")
        self.file_path_edit.setReadOnly(True)
        file_layout.addWidget(self.file_path_edit, 0, 1)
        
        browse_btn = QPushButton("浏览文件")
        browse_btn.clicked.connect(self.select_file)
        file_layout.addWidget(browse_btn, 0, 2)
        
        # 转换信息提示
        self.convert_info_label = QLabel("⏳ 未选择文件")
        self.convert_info_label.setWordWrap(True)
        self.convert_info_label.setStyleSheet("color: #666; font-size: 9pt; padding: 8px; background-color: #f5f5f5; border-radius: 4px;")
        file_layout.addWidget(self.convert_info_label, 1, 0, 1, 3)
        
        main_layout.addWidget(file_group)
        
        # ============ 处理进度区域 ============
        progress_group = QGroupBox("2. 分析进度")
        progress_layout = QVBoxLayout(progress_group)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        progress_layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel("就绪")
        self.status_label.setStyleSheet("font-size: 10pt; color: #666; padding: 5px;")
        progress_layout.addWidget(self.status_label)
        
        main_layout.addWidget(progress_group)
        
        # ============ CPK失败记录汇总表格 ============
        result_group = QGroupBox("3. CPK失败记录汇总")
        result_layout = QVBoxLayout(result_group)
        
        # 统计信息标签
        self.stats_label = QLabel("共 0 条失败记录")
        self.stats_label.setStyleSheet("color: #1976D2; font-size: 10pt; font-weight: bold; padding: 5px;")
        result_layout.addWidget(self.stats_label)
        
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(8)  # 增加一列No.
        self.result_table.setHorizontalHeaderLabels(['Date', 'No.', 'Cav No.', 'SPC', 'FAI#', 'TargetCPK', 'CalCPK', 'Result'])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)  # SPC列
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.result_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        result_layout.addWidget(self.result_table)
        
        main_layout.addWidget(result_group)
        
        # ============ 详细处理日志 ============
        log_group = QGroupBox("4. 详细处理日志")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #1e1e1e;
                color: #d4d4d4;
                font-family: Consolas, 'Courier New', monospace;
                font-size: 10pt;
                border: none;
            }
        """)
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        
        main_layout.addWidget(log_group)
        
        # ============ 按钮区域 ============
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        clear_log_btn = QPushButton("清空日志")
        clear_log_btn.setObjectName("danger")
        clear_log_btn.clicked.connect(self.clear_log)
        button_layout.addWidget(clear_log_btn)
        
        reset_btn = QPushButton("重置")
        reset_btn.setObjectName("warning")
        reset_btn.clicked.connect(self.reset)
        button_layout.addWidget(reset_btn)
        
        self.analyze_btn = QPushButton("开始分析")
        self.analyze_btn.setObjectName("success")
        self.analyze_btn.setMinimumWidth(150)
        self.analyze_btn.setStyleSheet("""
            QPushButton {
                background-color: #2E7D32;
                font-size: 12pt;
                padding: 10px 25px;
            }
        """)
        self.analyze_btn.clicked.connect(self.start_analysis)
        button_layout.addWidget(self.analyze_btn)
        
        main_layout.addLayout(button_layout)
        
        # 初始化日志
        self.log("=" * 80)
        self.log("CPK失败记录分析工具 v4.2")
        self.log("=" * 80)
        self.log("▶ 核心功能:")
        self.log("  1. 自动将.xls文件在同目录另存为.xlsx（仅保留数值）")
        self.log("  2. 不修改原始.xls文件")
        self.log("  3. 仅分析以CAV开头的sheet")
        self.log("  4. 比较E列TargetCPK和X列Cpk")
        self.log("  5. 在原文件最后添加CPK_FAIL_Summary汇总表")
        self.log("  6. 新增No.列，对应A列序号")
        self.log("=" * 80)
    
    def select_file(self):
        """选择文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择CPK报告文件",
            "",
            "Excel文件 (*.xls *.xlsx);;旧版Excel (*.xls);;新版Excel (*.xlsx);;所有文件 (*.*)"
        )
        
        if file_path:
            self.current_file_path = file_path
            self.file_path_edit.setText(file_path)
            self.log(f"已选择文件: {os.path.basename(file_path)}")
            
            if file_path.lower().endswith('.xls'):
                file_dir = os.path.dirname(file_path)
                file_name = os.path.splitext(os.path.basename(file_path))[0]
                converted_name = f"{file_name}_converted.xlsx"
                
                info_text = f"🔄 检测到 .xls 格式 | 将在同目录另存为: {converted_name} | 注意：转换会丢失公式"
                self.convert_info_label.setText(info_text)
                self.convert_info_label.setStyleSheet("color: #b26a00; font-size: 9pt; padding: 8px; background-color: #FFF3E0; border-radius: 4px; border-left: 4px solid #FF9800;")
            else:
                self.convert_info_label.setText(
                    f"✓ 已选择 .xlsx 格式文件，可直接分析"
                )
                self.convert_info_label.setStyleSheet("color: #2E7D32; font-size: 9pt; padding: 8px; background-color: #E8F5E9; border-radius: 4px; border-left: 4px solid #4CAF50;")
    
    def start_analysis(self):
        """开始分析"""
        if not self.current_file_path:
            QMessageBox.warning(self, "警告", "请先选择文件！")
            return
        
        if not os.path.exists(self.current_file_path):
            QMessageBox.critical(self, "错误", "文件不存在！")
            return
        
        # 禁用开始按钮
        self.analyze_btn.setEnabled(False)
        self.analyze_btn.setText("分析中...")
        
        # 清空表格
        self.result_table.setRowCount(0)
        self.stats_label.setText("分析中...")
        
        # 启动分析线程
        self.analysis_thread = AnalysisThread(self.current_file_path)
        
        # 连接信号
        self.analysis_thread.log_signal.connect(self.log)
        self.analysis_thread.progress_signal.connect(self.progress_bar.setValue)
        self.analysis_thread.status_signal.connect(self.status_label.setText)
        self.analysis_thread.error_signal.connect(self.show_error)
        self.analysis_thread.result_signal.connect(self.update_result_table)
        self.analysis_thread.finished_signal.connect(self.analysis_finished)
        
        # 启动线程
        self.analysis_thread.start()
        
        self.log("\n▶▶▶ 开始分析任务...")
        self.status_label.setText("正在分析中...")
        self.progress_bar.setValue(0)
    
    def update_result_table(self, failure_data):
        """更新结果表格"""
        self.failure_data = failure_data
        
        # 更新统计信息
        self.stats_label.setText(f"共 {len(failure_data)} 条失败记录")
        
        # 设置表格行数
        self.result_table.setRowCount(len(failure_data))
        
        # 填充数据
        for row, record in enumerate(failure_data):
            # Date
            date_item = QTableWidgetItem(record['date'])
            date_item.setTextAlignment(Qt.AlignCenter)
            self.result_table.setItem(row, 0, date_item)
            
            # No.
            no_item = QTableWidgetItem(record['no'])
            no_item.setTextAlignment(Qt.AlignCenter)
            self.result_table.setItem(row, 1, no_item)
            
            # Cav No.
            cav_item = QTableWidgetItem(record['cav_no'])
            cav_item.setTextAlignment(Qt.AlignCenter)
            self.result_table.setItem(row, 2, cav_item)
            
            # SPC
            spc_item = QTableWidgetItem(record['spc'])
            self.result_table.setItem(row, 3, spc_item)
            
            # FAI#
            fai_item = QTableWidgetItem(record['fai'])
            fai_item.setTextAlignment(Qt.AlignCenter)
            self.result_table.setItem(row, 4, fai_item)
            
            # TargetCPK
            target_item = QTableWidgetItem(f"{record['target_cpk']:.2f}")
            target_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.result_table.setItem(row, 5, target_item)
            
            # CalCPK
            cal_item = QTableWidgetItem(f"{record['cal_cpk']:.6f}")
            cal_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.result_table.setItem(row, 6, cal_item)
            
            # Result
            result_item = QTableWidgetItem("FAIL")
            result_item.setBackground(QBrush(QColor(192, 0, 0)))
            result_item.setForeground(QBrush(QColor(255, 255, 255)))
            result_item.setTextAlignment(Qt.AlignCenter)
            result_item.setFont(QFont("Microsoft YaHei", 9, QFont.Bold))
            self.result_table.setItem(row, 7, result_item)
        
        # 调整列宽
        self.result_table.resizeColumnsToContents()
        self.result_table.setColumnWidth(3, 400)  # SPC列宽一些
    
    def analysis_finished(self, success, message, failure_count):
        """分析完成"""
        # 恢复按钮
        self.analyze_btn.setEnabled(True)
        self.analyze_btn.setText("开始分析")
        
        if success:
            self.status_label.setText(f"分析完成")
            self.progress_bar.setValue(100)
            
            if failure_count > 0:
                msg = f"✅ 分析完成！\n\n发现 {failure_count} 条CPK失败记录\n\n"
                msg += "📊 输出文件：\n"
                
                if self.current_file_path.lower().endswith('.xls'):
                    file_dir = os.path.dirname(self.current_file_path)
                    file_name = os.path.splitext(os.path.basename(self.current_file_path))[0]
                    converted_name = f"{file_name}_converted.xlsx"
                    msg += f"• 已添加到: {converted_name}\n"
                else:
                    msg += f"• 已添加到: {os.path.basename(self.current_file_path)}\n"
                
                msg += "• 工作表位置: 最后\n"
                msg += "• 工作表名称: CPK_FAIL_Summary\n"
                msg += "• 新增列: No. (对应A列序号)\n"
                msg += "• 日期格式: YYYY-MM-DD\n"
                msg += "• 失败记录: 红色背景"
                
                QMessageBox.information(self, "分析完成", msg)
            else:
                QMessageBox.information(self, "分析完成", "✅ 分析完成！\n\n未发现CPK失败记录")
        else:
            self.status_label.setText("分析失败")
            self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: #d32f2f; }")
            QMessageBox.critical(self, "错误", f"❌ 处理失败:\n{message}")
    
    def show_error(self, error_msg):
        """显示错误"""
        self.log(f"❌ 错误: {error_msg}")
    
    def log(self, message):
        """添加日志"""
        timestamp = datetime.now().strftime("[%H:%M:%S]")
        self.log_text.append(f"{timestamp} {message}")
        
        # 自动滚动到底部
        cursor = self.log_text.textCursor()
        cursor.movePosition(QTextCursor.End)
        self.log_text.setTextCursor(cursor)
    
    def clear_log(self):
        """清空日志"""
        self.log_text.clear()
        self.log("日志已清空")
    
    def reset(self):
        """重置"""
        self.current_file_path = None
        self.converted_file_path = None
        self.failure_data = []
        self.file_path_edit.clear()
        self.progress_bar.setValue(0)
        self.status_label.setText("就绪")
        self.result_table.setRowCount(0)
        self.stats_label.setText("共 0 条失败记录")
        
        self.convert_info_label.setText("⏳ 未选择文件")
        self.convert_info_label.setStyleSheet("color: #666; font-size: 9pt; padding: 8px; background-color: #f5f5f5; border-radius: 4px;")
        
        self.clear_log()
        self.log("已重置")


def main():
    """主函数"""
    app = QApplication(sys.argv)
    app.setFont(QFont("Microsoft YaHei", 9))
    
    window = CPKAnalyzerWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()